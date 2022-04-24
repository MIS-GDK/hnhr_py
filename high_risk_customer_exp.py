import xlwt
import openpyxl
import cx_Oracle


def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            if value[i][j] is None:
                invalue = ""
            else:
                invalue = str(value[i][j])
            sheet.cell(row=i + 1, column=j + 1, value=invalue)
    sheet.insert_rows(1)
    sheet.cell(row=1, column=1, value="独立单元ID")
    sheet.cell(row=1, column=2, value="独立单元名称")
    sheet.cell(row=1, column=3, value="客户ID")
    sheet.cell(row=1, column=4, value="客户名称")
    sheet.cell(row=1, column=5, value="收单方ID")
    sheet.cell(row=1, column=6, value="收单方名称")
    sheet.cell(row=1, column=7, value="收单方编码")
    sheet.cell(row=1, column=8, value="业务员ID")
    sheet.cell(row=1, column=9, value="业务员名称")
    sheet.cell(row=1, column=10, value="一级部门ID")
    sheet.cell(row=1, column=11, value="一级部门名称")
    sheet.cell(row=1, column=12, value="二级部门ID")
    sheet.cell(row=1, column=13, value="二级部门名称")
    sheet.cell(row=1, column=14, value="客户分类")
    sheet.cell(row=1, column=15, value="客户分类名称")
    sheet.cell(row=1, column=16, value="业态")
    sheet.cell(row=1, column=17, value="账期方式")
    sheet.cell(row=1, column=18, value="应收余额")
    sheet.cell(row=1, column=19, value="逾期金额")
    sheet.cell(row=1, column=20, value="欠款天数")
    sheet.cell(row=1, column=21, value="账期天数")
    sheet.cell(row=1, column=22, value="超期天数")
    sheet.cell(row=1, column=23, value="截止当日未回款天数")
    sheet.cell(row=1, column=24, value="客户风险等级")
    sheet.cell(row=1, column=25, value="客户风险类别")
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")


conn = cx_Oracle.connect("hrhnprod/9bcPa4hr16HN@192.168.0.43:1525/HRHNDB")

cursor = conn.cursor()
cursor.execute("SELECT * FROM High_Risk_Customer_Sorts_Tl")
list1 = cursor.fetchall()

write_excel_xlsx("C:/Users/Administrator/Desktop/高风险客户20210715.xlsx", "高风险客户", list1)
