import pyodbc
import xlwt
import openpyxl


def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
    sheet.insert_rows(1)
    sheet.cell(row=1,column=1,value='人员')
    sheet.cell(row=1,column=2,value='部门')
    sheet.cell(row=1,column=3,value='是否有效')
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")


print(pyodbc.drivers())
conn = pyodbc.connect(
    "DRIVER={SQL Server};SERVER=192.168.0.16;DATABASE=ysdata;UID=ysadmin;PWD=0wM71IakpU&1s&"
)

cursor = conn.cursor()
cursor.execute("select RTRIM(dzyname),RTRIM(zyzw),RTRIM(beactive) from zhiydoc")
list1 = cursor.fetchall()

write_excel_xlsx(
    "C:/Users/Administrator/Desktop/预算名单20210609.xlsx", "预算名单", list1
)


