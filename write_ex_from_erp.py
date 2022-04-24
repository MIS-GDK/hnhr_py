import cx_Oracle
import xlwt
import openpyxl

v_count = 1
rownumber = 0


def write_excel_xlsx(path, sheet_name, value):
    global v_count
    global rownumber
    index = len(value)
    # workbook = openpyxl.Workbook()
    # sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1 + rownumber, column=j + 1, value=str(value[i][j]))

    rownumber = rownumber + index
    if v_count == 1:
        sheet.insert_rows(1)
        sheet.cell(row=1, column=1, value="业务日期")
        sheet.cell(row=1, column=2, value="客户名称")
        sheet.cell(row=1, column=3, value="通用名")
        sheet.cell(row=1, column=4, value="规格")
        sheet.cell(row=1, column=5, value="单位")
        sheet.cell(row=1, column=6, value="数量")
        sheet.cell(row=1, column=7, value="含税单价")
        sheet.cell(row=1, column=8, value="账期方式")
        sheet.cell(row=1, column=9, value="单据类型")
        sheet.cell(row=1, column=10, value="毛利额")
        sheet.cell(row=1, column=11, value="毛利率")
        sheet.cell(row=1, column=12, value="生产厂家")
        sheet.cell(row=1, column=13, value="税率")
        sheet.cell(row=1, column=14, value="税额")
        sheet.cell(row=1, column=15, value="无税金额")
        sheet.cell(row=1, column=16, value="出库日期")
        sheet.cell(row=1, column=17, value="不含税进价")
        sheet.cell(row=1, column=18, value="不含税进货金额")
        sheet.cell(row=1, column=19, value="含税进价")
        sheet.cell(row=1, column=20, value="含税进货金额")
        sheet.cell(row=1, column=21, value="含税金额")
        sheet.cell(row=1, column=22, value="供应商")
        sheet.cell(row=1, column=23, value="批号")
        sheet.cell(row=1, column=24, value="客户类型")
        sheet.cell(row=1, column=25, value="未回款数量")
        sheet.cell(row=1, column=26, value="未回款金额")
    workbook.save(path)
    v_count = v_count + 1
    # print("xlsx格式表格写入数据成功！")


conn = cx_Oracle.connect("hrhnprod/9bcPa4hr16HN@192.168.0.43:1525/HRHNDB")

cursor = conn.cursor()
cursor.execute(
    "SELECT a.Credate, "
    + "a.Customname, "
    + "a.Goodsname, "
    + "a.Goodstype, "
    + "a.Goodsunit, "
    + "a.Goodsqty, "
    + "a.Unitprice, "
    + "(SELECT Sdd.Ddlname "
    + "FROM Sys_Ddl_Dtl_v Sdd "
    + "WHERE a.Accountway = Sdd.Ddlid "
    + "AND Sdd.Sysid = 101471), "
    + "Decode(a.Settype, 1, '销售出库', 2, '销售退回', 3, '销售折扣'), "
    + "a.Gross, "
    + "a.Grossrate, "
    + "a.Factoryname, "
    + "a.Taxrate, "
    + "a.Taxmoney, "
    + "a.Notaxmoney, "
    + "a.Outdate, "
    + "a.Notaxsuprice, "
    + "a.Notaxsu_Total, "
    + "a.Suunitprice, "
    + "a.Su_Total, "
    + "a.Hsje, "
    + "a.Supplyname, "
    + "a.Lotno, "
    + "(SELECT Sdd.Ddlname "
    + "FROM Sys_Ddl_Dtl_v Sdd "
    + "WHERE a.Customertype = Sdd.Ddlid "
    + "AND Sdd.Sysid = 781), "
    + "a.Norecsqty, "
    + "a.Norecmoney "
    + "FROM Bms_Sa_Sales_All_v a "
    + "WHERE a.Credate >= To_Date('20190101', 'yyyymmdd') "
    + "AND a.Credate < To_Date('20200101', 'yyyymmdd') "
    + "AND a.Entryid = 9 and rownum <= 500000 "
    + "ORDER BY a.Credate"
)
# list1 = cursor.fetchall()

# write_excel_xlsx("C:/Users/Administrator/Desktop/预算名单20210727.xlsx", "销售明细2019", list1)
# res = self.cr.fetchmany(1000)

workbook = openpyxl.Workbook()
sheet = workbook.active
while True:
    list1 = cursor.fetchmany(10000)
    if not list1:
        break
    else:
        write_excel_xlsx(
            "C:/Users/Administrator/Desktop/销售明细2019.xlsx", "销售明细2019", list1
        )
