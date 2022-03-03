from openpyxl import load_workbook


wb = load_workbook(r"C:\Users\Administrator\Desktop\1.xlsx")
# 1、获取sheet名称
name_list = wb.sheetnames
print(name_list)

# for sheet_object in wb:
#     print(sheet_object.)
sheet = wb.worksheets[0]
# print(sheet.rows)
# for row in sheet.rows:
#     print(row)
# 按行读取
for row in sheet.iter_rows(min_row=1, max_row=3):
    print(row)
